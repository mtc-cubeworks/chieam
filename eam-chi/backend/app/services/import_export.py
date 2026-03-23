from __future__ import annotations

import csv
import io
from typing import Any, Iterable
from urllib.parse import urlparse, parse_qs
from urllib.request import urlopen

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.meta.registry import EntityMeta, MetaRegistry
from app.infrastructure.database.repositories.entity_repository import get_entity_model
from app.services.naming import NamingService


SYSTEM_FIELDS = {"id", "created_at", "updated_at"}


def _safe_display_field(meta: EntityMeta | None, model: Any) -> str:
    """Return a best-effort display field that exists on the SQLAlchemy model."""
    preferred = []
    if meta and getattr(meta, "title_field", None):
        preferred.append(meta.title_field)
    preferred.extend(
        [
            "name",
            "item_name",
            "organization_name",
            "site_name",
            "department_name",
            "vendor_name",
            "company_name",
            "model_name",
            "trade_name",
            "asset_tag",
            "description",
        ]
    )
    preferred.append("id")
    for field in preferred:
        if hasattr(model, field):
            return field
    return "id"


def get_import_fields(meta: EntityMeta) -> list[str]:
    # Allow importing `id` optionally:
    # - If provided, it will be used as-is.
    # - If not provided, NamingService will generate it if naming is enabled.
    fields = [
        f.name
        for f in meta.fields
        if not f.hidden and not f.readonly and f.name not in {"created_at", "updated_at"}
    ]
    if "id" in fields:
        fields.remove("id")
        fields.insert(0, "id")
    return fields


def _normalize_header(value: str) -> str:
    return value.strip()


def parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [_normalize_header(h or "") for h in (reader.fieldnames or [])]
    rows: list[dict[str, Any]] = []
    for row in reader:
        normalized = {
            _normalize_header(k): (v.strip() if isinstance(v, str) else v)
            for k, v in row.items()
        }
        rows.append(normalized)
    return headers, rows


def parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active
    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx == 0:
            headers = [_normalize_header(str(cell or "")) for cell in row]
            continue
        record = {}
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                continue
            record[headers[col_idx]] = cell
        rows.append(record)
    return headers, rows


def parse_upload(file: UploadFile) -> tuple[list[str], list[dict[str, Any]]]:
    content = file.file.read()
    if file.filename and file.filename.lower().endswith(".csv"):
        return parse_csv(content)
    if file.filename and file.filename.lower().endswith(".xlsx"):
        return parse_xlsx(content)
    raise ValueError("Unsupported file type. Please upload a CSV or XLSX file.")


def parse_google_sheet(url: str) -> tuple[list[str], list[dict[str, Any]]]:
    parsed = urlparse(url)
    if "docs.google.com" not in parsed.netloc:
        raise ValueError("Invalid Google Sheets URL")
    parts = parsed.path.split("/")
    if "spreadsheets" not in parts:
        raise ValueError("Invalid Google Sheets URL")
    sheet_id = parts[parts.index("d") + 1] if "d" in parts else None
    if not sheet_id:
        raise ValueError("Invalid Google Sheets URL")

    qs = parse_qs(parsed.query)
    gid = qs.get("gid", ["0"])[0]
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    with urlopen(export_url) as response:
        content = response.read()
    return parse_csv(content)


def build_template(meta: EntityMeta, link_options: dict[str, list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    fields = get_import_fields(meta)
    ws.append(fields)

    # Create References sheet with index-based legends
    ref = wb.create_sheet("References")
    ref.append(["Field", "Index", "Value"])
    
    for field_name, values in link_options.items():
        # Add header row for this field
        ref.append([field_name, "", ""])
        # Add indexed values (0, 1, 2, ...)
        for idx, value in enumerate(values[:500]):  # Limit to 500 entries
            ref.append(["", str(idx), value])
        # Add blank row between fields
        ref.append(["", "", ""])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def get_link_fields(meta: EntityMeta) -> list[str]:
    return [f.name for f in meta.fields if f.field_type == "link" and f.link_entity]


def validate_headers(meta: EntityMeta, headers: list[str]) -> list[str]:
    expected = set(get_import_fields(meta))
    provided = set(headers)
    # `id` is optional on import; created_at/updated_at are ignored if provided.
    missing = (expected - {"id"}) - provided
    extra = provided - expected
    extra = extra - {"created_at", "updated_at"}
    errors: list[str] = []
    if missing:
        errors.append(f"Missing columns: {', '.join(sorted(missing))}")
    if extra:
        errors.append(f"Unexpected columns: {', '.join(sorted(extra))}")
    return errors


async def build_link_options(db: AsyncSession, meta: EntityMeta) -> dict[str, list[str]]:
    """Build link options with display values for template generation."""
    options: dict[str, list[str]] = {}
    for field in meta.fields:
        if field.field_type != "link" or not field.link_entity:
            continue
        linked_meta = MetaRegistry.get(field.link_entity)
        linked_model = get_entity_model(field.link_entity)
        if not linked_meta or not linked_model:
            continue
        display_field = _safe_display_field(linked_meta, linked_model)
        stmt = select(getattr(linked_model, display_field)).order_by(getattr(linked_model, display_field))
        rows = (await db.execute(stmt)).scalars().all()
        options[field.name] = [str(v) for v in rows if v is not None]
    return options


async def build_link_mappings(db: AsyncSession, meta: EntityMeta) -> dict[str, dict[str, str]]:
    """Build link field mappings: {field_name: {display_value: id}}"""
    mappings: dict[str, dict[str, str]] = {}
    for field in meta.fields:
        if field.field_type != "link" or not field.link_entity:
            continue
        linked_meta = MetaRegistry.get(field.link_entity)
        linked_model = get_entity_model(field.link_entity)
        if not linked_meta or not linked_model:
            continue
        display_field = _safe_display_field(linked_meta, linked_model)
        stmt = select(linked_model).order_by(getattr(linked_model, display_field))
        rows = (await db.execute(stmt)).scalars().all()
        
        field_mapping = {}
        for row in rows:
            display_value = str(getattr(row, display_field, None) or getattr(row, "id"))
            field_mapping[display_value] = str(row.id)
        
        mappings[field.name] = field_mapping
    return mappings


async def resolve_link_value(db: AsyncSession, link_entity: str, value: Any, link_mappings: dict[str, str] = None) -> str | None:
    """
    Resolve link field value to ID.
    Supports:
    1. Numeric index (0, 1, 2) - maps to ID using legend order
    2. Display value (e.g., "Site A") - looks up by title field
    3. Direct ID - returns as-is if valid
    """
    if value is None or value == "":
        return None
    
    value_str = str(value).strip()

    # Fast path: treat as ID if it exists
    linked_model = get_entity_model(link_entity)
    if linked_model:
        res = await db.execute(select(linked_model).where(getattr(linked_model, "id") == value_str))
        if res.scalar_one_or_none():
            return value_str
    
    # Check if it's a numeric index
    if value_str.isdigit():
        index = int(value_str)
        # Get all records ordered by title field (same order as legend)
        linked_meta = MetaRegistry.get(link_entity)
        linked_model = get_entity_model(link_entity)
        if not linked_meta or not linked_model:
            return None
        
        display_field = _safe_display_field(linked_meta, linked_model)
        stmt = select(linked_model).order_by(getattr(linked_model, display_field)).offset(index).limit(1)
        result = (await db.execute(stmt)).scalar_one_or_none()
        if result:
            return str(result.id)
    
    # Try using pre-built mapping if provided
    if link_mappings and value_str in link_mappings:
        return link_mappings[value_str]
    
    # Fallback: lookup by title field
    linked_meta = MetaRegistry.get(link_entity)
    linked_model = get_entity_model(link_entity)
    if not linked_meta or not linked_model:
        return None
    display_field = _safe_display_field(linked_meta, linked_model)
    if display_field == "id":
        return None
    stmt = select(linked_model).where(getattr(linked_model, display_field) == value_str)
    result = (await db.execute(stmt)).scalar_one_or_none()
    if not result:
        return None
    return str(result.id)


def convert_field_value(value: Any, field_name: str) -> Any:
    """Convert field value to appropriate type for database storage."""
    if value is None or value == "":
        return None
    
    # If it's already a string and not numeric, return as-is
    if isinstance(value, str):
        # Check if it's a numeric string that should stay as string (like serial numbers)
        if field_name in ['serial_number', 'tag_number', 'phone', 'code', 'identifier']:
            return value
        # Try to parse as number for numeric fields
        if value.replace('.', '').replace('-', '').isdigit():
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                return value
        return value
    
    # Convert numbers to strings for specific fields
    if isinstance(value, (int, float)):
        if field_name in ['serial_number', 'tag_number', 'description', 'notes', 'name', 'code']:
            return str(value)
        return value
    
    return str(value)


async def validate_rows(
    db: AsyncSession,
    meta: EntityMeta,
    rows: list[dict[str, Any]],
    mode: str = "create",
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []
    validated: list[dict[str, Any]] = []
    model = get_entity_model(meta.name)

    link_fields = [f for f in meta.fields if f.field_type == "link" and f.link_entity]
    # Build mappings once per validation pass (performance + stable resolution)
    link_mappings = await build_link_mappings(db, meta) if link_fields else {}

    for idx, row in enumerate(rows, start=2):
        row_errors = []
        data: dict[str, Any] = {}
        record_id = row.get("id")

        if mode == "update":
            if record_id is None or str(record_id).strip() == "":
                row_errors.append({
                    "field": "id",
                    "message": "Record id is required for update mode",
                })
            elif not model:
                row_errors.append({
                    "field": "id",
                    "message": "Model not found",
                })
            else:
                existing = await db.execute(select(model).where(getattr(model, "id") == str(record_id).strip()))
                if existing.scalar_one_or_none() is None:
                    row_errors.append({
                        "field": "id",
                        "message": f"Record '{record_id}' does not exist",
                    })

        for field in meta.fields:
            if field.hidden or field.readonly or field.name in {"created_at", "updated_at"}:
                continue
            value = row.get(field.name)
            if mode == "update" and field.name != "id" and (value is None or value == ""):
                continue
            if field.required and (value is None or value == ""):
                row_errors.append({
                    "field": field.name,
                    "message": "Required field is missing",
                })
                continue
            if field.field_type == "link" and field.link_entity:
                resolved = await resolve_link_value(
                    db,
                    field.link_entity,
                    value,
                    link_mappings=link_mappings.get(field.name),
                )
                if value and not resolved:
                    row_errors.append({
                        "field": field.name,
                        "message": f"Invalid link value '{value}'",
                    })
                data[field.name] = resolved
            else:
                # Convert value to appropriate type
                converted_value = convert_field_value(value, field.name)
                data[field.name] = converted_value

        if row_errors:
            errors.append({"row": idx, "errors": row_errors})
        else:
            validated.append(data)

    return validated, errors


async def create_records(
    db: AsyncSession,
    meta: EntityMeta,
    records: list[dict[str, Any]],
    user: Any = None,
) -> tuple[int, list[dict[str, Any]]]:
    """Create records with auto-generated IDs and duplicate checking."""
    model = get_entity_model(meta.name)
    if not model:
        raise ValueError("Model not found")
    
    created_count = 0
    duplicates: list[dict[str, Any]] = []
    
    # Get the title field for duplicate checking
    title_field = meta.title_field
    
    for record in records:
        # Check for duplicates based on title field
        if title_field and title_field in record:
            stmt = select(model).where(getattr(model, title_field) == record[title_field])
            existing = (await db.execute(stmt)).scalar_one_or_none()
            if existing:
                duplicates.append(record)
                continue
        
        # Determine record id
        record_data = record.copy()
        provided_id = record_data.get("id")
        if provided_id is not None and str(provided_id).strip() == "":
            record_data.pop("id", None)
            provided_id = None

        if not provided_id:
            if meta.naming and getattr(meta.naming, "enabled", False):
                record_data["id"] = await NamingService.generate_id(db, meta.naming)
            else:
                raise ValueError(f"Cannot import '{meta.name}' without 'id' because naming is disabled")

        # Inject created_by from current user
        if user and hasattr(model, "created_by"):
            record_data.setdefault("created_by", getattr(user, "id", None))

        # Create new record
        try:
            obj = model(**record_data)
            db.add(obj)
            await db.flush()  # Get the ID without committing
            created_count += 1
        except Exception as e:
            await db.rollback()
            raise e
    
    await db.commit()
    return created_count, duplicates


async def update_records(
    db: AsyncSession,
    meta: EntityMeta,
    records: list[dict[str, Any]],
    user: Any = None,
) -> tuple[int, list[dict[str, Any]]]:
    model = get_entity_model(meta.name)
    if not model:
        raise ValueError("Model not found")

    # Build scope filter once for all records
    scope_clause = None
    if user:
        from app.services.rbac import RBACService
        scope_clause = RBACService.build_scope_filter(user, model)

    updated_count = 0
    missing: list[dict[str, Any]] = []

    try:
        for record in records:
            record_id = record.get("id")
            if record_id is None or str(record_id).strip() == "":
                missing.append(record)
                continue

            stmt = select(model).where(getattr(model, "id") == str(record_id).strip())
            if scope_clause is not None:
                stmt = stmt.where(scope_clause)
            existing = await db.execute(stmt)
            obj = existing.scalar_one_or_none()
            if obj is None:
                missing.append(record)
                continue

            for key, value in record.items():
                if key in {"id", "created_at", "updated_at"}:
                    continue
                if value is None:
                    continue
                setattr(obj, key, value)

            # Inject modified_by from current user
            if user and hasattr(model, "modified_by"):
                obj.modified_by = getattr(user, "id", None)

            await db.flush()
            updated_count += 1

        await db.commit()
    except Exception:
        await db.rollback()
        raise

    return updated_count, missing


# Re-export from export_service for backward compatibility
from app.services.export_service import export_records, get_export_fields  # noqa: F811,E402
